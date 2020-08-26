# -*- coding: utf-8 -*-
"""
@Time    : 2020/2/15 下午2:21

@File    : test_rpsIndex.py

@author  : pchaos
@license : Copyright(C), pchaos
@Contact : p19992003#gmail.com
"""
import unittest
import datetime
import time
import pandas as pd
import os
import matplotlib.pyplot as plt
import pyperclip
import numpy as np
import QUANTAXIS as qa
from tabulate import tabulate
from userFunc import cal_ret, get_RPS
from userFunc import full_pickle, loosen_pickle, compressed_pickle, decompress_pickle
from userFunc import RPSIndex
from userFunc import read_zxg
from userFunc import codeInfo
from testing.qaHelper.fetcher.qhtestbase import QhBaseTestCase


def tableize(df):
    if not isinstance(df, pd.DataFrame):
        return
    df_columns = df.columns.tolist()
    max_len_in_lst = lambda lst: len(sorted(lst, reverse=True, key=len)[0])
    align_center = lambda st, sz: "{0}{1}{0}".format(" " * (1 + (sz - len(st)) // 2), st)[:sz] if len(st) < sz else st
    align_right = lambda st, sz: "{0}{1} ".format(" " * (sz - len(st) - 1), st) if len(st) < sz else st
    max_col_len = max_len_in_lst(df_columns)
    max_val_len_for_col = dict(
        [(col, max_len_in_lst(df.iloc[:, idx].astype('str'))) for idx, col in enumerate(df_columns)])
    col_sizes = dict([(col, 2 + max(max_val_len_for_col.get(col, 0), max_col_len)) for col in df_columns])
    build_hline = lambda row: '+'.join(['-' * col_sizes[col] for col in row]).join(['+', '+'])
    build_data = lambda row, align: "|".join(
        [align(str(val), col_sizes[df_columns[idx]]) for idx, val in enumerate(row)]).join(['|', '|'])
    hline = build_hline(df_columns)
    out = [hline, build_data(df_columns, align_center), hline]
    for _, row in df.iterrows():
        out.append(build_data(row.tolist(), align_right))
    out.append(hline)
    return "\n".join(out)


class TestRPSIndex(QhBaseTestCase):
    def setUp(self) -> None:
        # 是否删除临时文件
        # self.deltmp = True
        self.deltmp = False

        fileName = '/tmp/indexCode.pickle'
        try:
            self.code = decompress_pickle(fileName)
        except Exception as e:
            self.code = list(qa.QA_fetch_index_list_adv()['code'][:])
            compressed_pickle(fileName, self.code)

        days = 365 * 1.2
        # days = 365 * 10
        self.start = datetime.datetime.now() - datetime.timedelta(days)
        self.end = datetime.datetime.now() - datetime.timedelta(0)
        fileName = '/tmp/data{}.pickle'.format(days)
        self.data2 = self.getIndexCalret(fileName, self.code, self.start, self.end)

    def tearDown(self) -> None:
        try:
            if self.deltmp:
                import os
                myCmd = 'rm /tmp/*.pbz2'
                os.system(myCmd)
        except Exception as e:
            pass

    def getIndexCalret(self, readFromfile, code, startDate, endDate):
        try:
            # 获取指数数据
            dataCalret = decompress_pickle(readFromfile)
        except Exception as e:
            data = qa.QA_fetch_index_day_adv(code, startDate, endDate)
            df = data.add_func(cal_ret)
            compressed_pickle(readFromfile, data.data)
            dataCalret = decompress_pickle(readFromfile)
            data2 = dataCalret
        return dataCalret

    def test_rps_class(self):
        """
        如果测试有错误，删除临时目录pickle文件
        """
        # 显示rps排名前10%的中文名称
        code = self.code
        rpsday = [20, 50]
        dfrps = self._getRPS(rpsday, self.data2)
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        rps = rpsIndex.rps()
        try:
            self.assertTrue(dfrps.sort_values(dfrps.columns[0], ascending=False).equals(rps),
                            "排序后不相等 {} {}".format(dfrps, rps))
        except:
            self.assertTrue(dfrps.equals(rps), "不相等 {} {}".format(dfrps, rps))
        # self.assertTrue(dfrps.equals(rps), "{} {}".format(dfrps.head(), rps.head()))
        print(rps.tail())
        print(rps.head(20))

    def test_rps_class_multi_rpsday(self):
        # 显示rps排名前10%的中文名称
        code = self.code
        rpsday = [20, 50, 120]
        dfrps = self._getRPS(rpsday, self.data2)
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        rps = rpsIndex.rps()
        try:
            self.assertTrue(dfrps.sort_values(dfrps.columns[0], ascending=False).equals(rps),
                            "排序后不相等 {} {}".format(dfrps, rps))
        except:
            self.assertTrue(dfrps.equals(rps), "不相等 {} {}".format(dfrps, rps))
        print(rps.tail())

    def test_rps_class_selectCode(self):
        # 显示rps排名前10%的ETF中文名称
        code = self.code
        rpsday = [20, 50, 89]
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        rps = rpsIndex.selectCode(code[0])
        print(rps.tail())
        rps.plot()
        rps.plot(x='RPS20', y='RPS{}'.format(rpsday[1]))
        rps.plot(x='RPS20', y='RPS{}'.format(rpsday[2]))
        plt.show()

    def _getRPS(self, rpsday, dataFrame):
        # data = qa.QA_DataStruct_Index_day(self.data2)
        data = qa.QA_DataStruct_Index_day(dataFrame)
        df = data.add_func(cal_ret, *rpsday)
        matching = [s for s in df.columns if "MARKUP" in s]
        self.assertTrue(len(matching) == len(rpsday), '计算周期不在返回的字段中')
        print(df.head())
        # 计算RPS
        dfg = df.groupby(level=0).apply(get_RPS, *rpsday)
        return dfg

    def test_rps_ETF_selectCode(self):
        # rps筛选特定code
        fn = 'zxgETF.txt'
        # code列表
        code = read_zxg(fn)
        rpsday = [20, 50]
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        rps = rpsIndex.selectCode(code)
        # print(rps.tail())
        # print(rps)
        self.assertTrue(len(rpsIndex.selectCode(code[:2])) == 2 * len(rpsIndex.selectCode(code[:1])), "两倍长度")

    def test_rps_ETFTOPN(self):
        # 显示rps排名前10%的ETF中文名称
        fn = 'zxgETF.txt'
        # code列表
        code = read_zxg(fn)
        rpsday = [20, 50]
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        n = 10
        rps = rpsIndex.rpsTopN(self.end, n)
        self.assertTrue(len(rps) > 0)

        print(rps)
        print("总数： {}".format(len(rps)))
        dfInfo = codeInfo(list(rps.index.levels[1]))
        print(dfInfo.name)

    def test_RPSIndex_rpsTopN2(self):
        # 显示rps排名前10%的ETF中文名称
        fn = 'zxgETF.txt'
        # code列表
        code = read_zxg(fn)
        rpsday = [20, 50]
        end = self.end
        # end = "2020-06-24"
        rpsIndex = RPSIndex(code, self.start, end, rpsday)
        n = 10
        rps = rpsIndex.rpsTopN2(startday=end, percentN=n)
        rps2 = rpsIndex.rpsTopN(end, n)
        self.assertTrue(len(rps) > 0)
        self.assertTrue(len(rps) == len(rps2), "长度不同 {} {}\n{} {}".format(len(rps), len(rps2), rps, rps2))
        try:
            self.assertTrue(rps.sort_values(rps.columns[0], ascending=False).equals(rps2),
                            "排序后不相等 {} {}".format(rps, rps2))
        except:
            self.assertTrue(rps.equals(rps2), "不相等 {} {}".format(rps, rps2))
        print(rps)
        print("总数： {}".format(len(rps)))
        dfInfo = codeInfo(list(rps.index.levels[1]))
        print(dfInfo.name)

    def test_RPSIndex_rpsTopN2_2(self):
        #  一段时间内，显示rps排名前10%的ETF中文名称
        fn = 'zxgETF.txt'
        # code列表
        code = read_zxg(fn)
        rpsday = [20, 50]
        rpsIndex = RPSIndex(code, self.start, self.end, rpsday)
        n = 10
        start = self.end - datetime.timedelta(10)
        rps = rpsIndex.rpsTopN2(startday=start, lastday=self.end, percentN=n)
        self.assertTrue(len(rps) > 0)
        # for day in reversed(rps.index.levels[0]):
        for day in rps.index.levels[0]:
            if day == pd.to_datetime(datetime.date(2020, 6, 30)):
                continue
            print("对比日期： {}".format(day))
            rps2 = rpsIndex.rpsTopN(day, n)
            df = rps.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            self.assertTrue(len(df) == len(rps2), "长度不同 {} {}\n{} {}".format(len(df), len(rps2), rps, rps2))
            try:
                obo = self.differOneByOne(df, rps2)
                self.assertTrue(len(obo) == 0, "{} {}".format(df, rps2))
                self.assertTrue(df.equals(rps2), "不相等 {} {}".format(df, rps2))
            except Exception as e:
                df = df.sort_values(by=df.columns[0], ascending=False, inplace=False)
                obo = self.differOneByOne(df, rps2)
                self.assertTrue(len(obo) == 0, "{} {}".format(df, rps2))
                self.assertTrue(df.equals(rps2), "不相等 {} {}".format(df, rps2))
        print(rps)
        print("总数： {}".format(len(rps)))
        dfInfo = codeInfo(list(rps.index.levels[1]))
        print(dfInfo.name)

    def test_readExcel_block10(self):
        """ 找出新进入强势区间的板块
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            print("新进入强势区间的板块")
            #
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            # df = pd.read_excel(filename, sheet_name=shedf['RPS10']> 95 & df['RPS10'].shift() < 95etName + "1", converters={'code': str})
            colname = df.columns[1]
            n = 90
            day = df.index.levels[0][-1]
            # 本日rps刚超过n
            dftop = df[(df[colname].groupby(level=1).shift() < n) & (df[colname] >= n)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            if len(dfp) > 0:
                pyperclip.copy(dfp.to_csv())
                pyperclip.paste()
                print("复制道粘贴板")
                print(dfp)
            else:
                print("no data")
            day = df.index.levels[0][-2]
            print("前一日rps强度")
            dfp2 = df.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp2['code'] = dfp2.code.apply(lambda x: str(x).zfill(6))
            dfp2.set_index(['date', 'code'], inplace=True)
            dfp2 = dfp2.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), codes), :]
            print(dfp2, "\n长度：", len(dfp2))
            self._nameDiff(dfp, dfp2)
            self.assertTrue(len(dfp) >= len(dfp2), "{} {}".format(len(dfp), len(dfp2)))

    def _nameDiff(self, dfp, dfp2):
        if len(dfp != len(dfp2)):
            test = list(np.setdiff1d(list(dfp.name), list(dfp2.name)))
            if len(test) > 0:
                print(f"上一天不存在：{test}")

    def test_readExcel2(self):
        """ 找出rps10新进入强势区间,rps20也在强势区间的板块
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            #
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            colname = df.columns[1]
            colname2 = df.columns[2]
            n = 87
            day = df.index.levels[0][-1]
            # 本日rps10刚超过n，并且rps20也超过n
            dftop = df[(df[colname].groupby(level=1).shift() < n) & (df[colname] >= n) & (df[colname2] >= n)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            print(dfp)
            day = df.index.levels[0][-2]
            print("前一日rps强度")
            dfp2 = df.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp2['code'] = dfp2.code.apply(lambda x: str(x).zfill(6))
            dfp2.set_index(['date', 'code'], inplace=True)
            dfp2 = dfp2.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), codes), :]
            print(dfp2, "\n长度：", len(dfp2))
            self.assertTrue(len(dfp) == len(dfp2), "{} {}".format(len(dfp), len(dfp2)))

    def test_readExcel3(self):
        """ 找出rps10新进入强势区间,rps20也在强势区间的板块
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            print("rps10新进入强势区间,rps20也在强势区间的板块")
            #
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            # df = pd.read_excel(filename, sheet_name=shedf['RPS10']> 95 & df['RPS10'].shift() < 95etName + "1", converters={'code': str})
            colname = df.columns[1]
            colname2 = df.columns[2]
            n, m = 87, 90
            day = df.index.levels[0][-1]
            # 本日rps10刚超过n，并且rps20也超过n
            dftop = df[(df[colname] >= n) & (df[colname2] >= m)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            print(dfp)
            day = df.index.levels[0][-2]
            print("前一日rps强度")
            dfp2 = df.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp2['code'] = dfp2.code.apply(lambda x: str(x).zfill(6))
            dfp2.set_index(['date', 'code'], inplace=True)
            dfp2 = dfp2.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), codes), :]
            print(dfp2, "\n长度：", len(dfp2))
            self.assertTrue(len(dfp) == len(dfp2), "{} {}".format(len(dfp), len(dfp2)))

    def test_readExcel4(self):
        """ 找出rps10新进入强势区间,rps20也在强势区间的板块
        本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            #
            print("本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n")
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            colname = df.columns[1]
            colname2 = df.columns[2]
            n = 87
            day = df.index.levels[0][-1]
            # 本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n
            dftop = df[(df[colname].groupby(level=1).shift(1) < n) & (df[colname] >= n) & (df[colname2] >= n * 0.8) & (
                    df[colname2].groupby(level=1).shift(1) < n)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            print(dfp)
            day = df.index.levels[0][-2]
            print(f"前一日rps强度({str(day)[:10]})")
            dfp2 = df.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :].reset_index(drop=False)
            dfp2['code'] = dfp2.code.apply(lambda x: str(x).zfill(6))
            dfp2.set_index(['date', 'code'], inplace=True)
            dfp2 = dfp2.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), codes), :]
            print(dfp2, f"\n长度：{len(dfp2)}, {len(dfp)}")
            self._nameDiff(dfp, dfp2)
            self._nameDiff(dfp2, dfp)
            self.assertTrue(len(dfp) >= len(dfp2), "{} {}".format(len(dfp), len(dfp2)))

    def test_readExcel_peroid(self):
        """ 找出rps10新进入强势区间,rps20也在强势区间的板块
        本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            #
            n = 87
            print(f'当日rps10刚超过{n}，rps20也超过{round(n * 0.8, 2)}, 前一日rps20小于{n}')
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            colname = df.columns[1]
            colname2 = df.columns[2]
            day = df.index.levels[0][-10]
            # 本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n
            dftop = df[(df[colname].groupby(level=1).shift(1) < n) & (df[colname] >= n) & (df[colname2] >= n * 0.8) & (
                    df[colname2].groupby(level=1).shift(1) < n)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(None))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            # print(dfp)
            print("满足条件个数")
            for day in df.index.levels[0][-10:]:
                print(f"{str(day)[:10]}  {len(dfp.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), slice(None)), :])}")
                # print(dfp.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), slice(None)), :])
            # print(tabulate(dfp, headers='keys', tablefmt='psql'))
            # print(tableize(dfp))
            #
            dfs = dfp.style.applymap(lambda x: "background-color: red" if x >= n else "background-color: white",
                                     subset=[colname2]) \
                .highlight_max(axis=1) \
                .format({colname: "${:10,.2f}"})

            self.df2xlsWithColor(dfs, filename)


    def test_readExcel_peroid_2(self):
        """ 找出rps20新进入强势区间,rps10也在强势区间的板块
        当日rps10刚超过n，并且rps20也超过n 前一日rps20小于n
        注意：产生rps文件时，要取前40%的数据
        """
        filename = '/tmp/rpstop.xlsx'
        sheetName = "rps"
        if os.path.exists(filename):
            #
            n = 87
            print(f'当日rps10刚超过{n}，rps20也超过{round(n, 2)}, 前一日rps20小于{n}')
            df = pd.read_excel(filename, sheet_name=sheetName, index_col=[0, 1], converters={'code': str})
            colname = df.columns[1]
            colname2 = df.columns[2]
            day = df.index.levels[0][-10]
            # 本日rps10刚超过n，并且rps20也超过n*0.8 前一日rps20小于n
            dftop = df[(df[colname2].groupby(level=1).shift(1) < n) & (df[colname] >= n) & (df[colname2] >= n)]
            # dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(day))), :]
            dfp = dftop.loc[(slice(pd.Timestamp(day), pd.Timestamp(None))), :].reset_index(drop=False)
            dfp['code'] = dfp.code.apply(lambda x: str(x).zfill(6))
            dfp.set_index(['date', 'code'], inplace=True)
            codes = dfp.index.levels[1]
            # print(dfp)
            print("满足条件个数")
            for day in df.index.levels[0][-10:]:
                print(f"{str(day)[:10]}  {len(dfp.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), slice(None)), :])}")
                # print(dfp.loc[(slice(pd.Timestamp(day), pd.Timestamp(day)), slice(None)), :])
            # print(tabulate(dfp, headers='keys', tablefmt='psql'))
            # print(tableize(dfp))
            #
            dfs = dfp.style.applymap(lambda x: "background-color: red" if x >= n else "background-color: white",
                                     subset=[colname2]) \
                .highlight_max(axis=1) \
                .format({colname: "${:10,.2f}"})

            self.df2xlsWithColor(dfs, filename)

    def df2xlsWithColor(self, dfs, filename):
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, NamedStyle
        from openpyxl.utils import get_column_letter
        # pass keyword args as dictionary
        writer_args = {
            'path': filename,
            'mode': 'a',
            'engine': 'openpyxl'}
        with pd.ExcelWriter(**writer_args) as writer:
            sheet_name = 'rpsColored'
            dfs.to_excel(writer, sheet_name)

            # set index column width
            ws = writer.sheets[sheet_name]
            value_cells = 'D2:{col}{row}'.format(
                col=get_column_letter(ws.max_column),
                row=ws.max_row)
            # for general styling, one has to iterate over
            # all cells individually
            for row in ws[value_cells]:
                for cell in row:
                    cell.number_format = '0.00'
            title_row = '1'
            ws.column_dimensions['A'].width = 21


if __name__ == '__main__':
    import subprocess

    list_files = subprocess.run(["ls", "-l"])
    print("The exit code was: %d" % list_files.returncode)

    unittest.main()
